# Copyright (c) 2025 Tadashi Suto


import os
import requests
import urllib3
import re
import threading
import customtkinter as ctk
import tkinter.filedialog as fd
import tkinter.messagebox as mb
from datetime import datetime
from PIL import Image
from openpyxl import load_workbook, Workbook

# Ignora avisos de verificação SSL
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Validação de e-mail
def email_valido(email):
    return re.match(r"[^@]+@[^@]+\.[^@]+", email)

# Validação de campo não vazio
def campo_valido(valor):
    return valor is not None and str(valor).strip() != ""

# Tela de Splash
class SplashScreen(ctk.CTk):
    def __init__(self):
        super().__init__()

        # Configuração da janela
        width, height = 500, 500
        screen_width, screen_height = self.winfo_screenwidth(), self.winfo_screenheight()
        x = (screen_width / 2) - (width / 2)
        y = (screen_height / 2) - (height / 2)
        self.geometry(f'{width}x{height}+{int(x)}+{int(y)}')
        self.overrideredirect(True)  # Remove a barra de título e bordas

        ctk.set_appearance_mode("dark")

        # Carrega e exibe o logo
        try:
            logo_image = ctk.CTkImage(Image.open("Designer.png"), size=(400, 400))
            logo_label = ctk.CTkLabel(self, image=logo_image, text="")
            logo_label.pack(pady=(40, 10), padx=20)
        except FileNotFoundError:
            logo_label = ctk.CTkLabel(self, text="Logo não encontrado", font=("Arial", 20))
            logo_label.pack(pady=(40, 10), padx=20)

        # Exibe a versão
        version_label = ctk.CTkLabel(self, text="Versão 1.0", font=("Arial", 14))
        version_label.pack(pady=10)

        # Agenda o fechamento do splash e a abertura do app principal
        self.after(3000, self.open_main_app)

    def open_main_app(self):
        self.destroy()
        App().mainloop()

# App principal
class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Emissor de Pesquisa - Amplifique.me")
        self.geometry("800x650")
        self.logs_excel = []

        ctk.set_appearance_mode("dark")
        ctk.set_default_color_theme("blue")

        self.emails_enviados_path = ctk.StringVar()
        self.pesquisa_path = ctk.StringVar()
        self.token = ctk.StringVar()
        self.expiration = ctk.StringVar(value="5") # Valor padrão

        self.criar_interface()

    def criar_interface(self):
        ctk.CTkLabel(self, text="Planilha de E-mails Já Enviados:").pack(pady=(10, 2))
        ctk.CTkEntry(self, textvariable=self.emails_enviados_path, width=500).pack()
        ctk.CTkButton(self, text="Selecionar", command=lambda: self.selecionar_arquivo(self.emails_enviados_path)).pack(pady=(2, 10))

        ctk.CTkLabel(self, text="Planilha da Pesquisa (dados a enviar):").pack(pady=(10, 2))
        ctk.CTkEntry(self, textvariable=self.pesquisa_path, width=500).pack()
        ctk.CTkButton(self, text="Selecionar", command=lambda: self.selecionar_arquivo(self.pesquisa_path)).pack(pady=(2, 10))

        ctk.CTkLabel(self, text="Token da Pesquisa:").pack(pady=(10, 2))
        ctk.CTkEntry(self, textvariable=self.token, show="*", width=500).pack()

        ctk.CTkLabel(self, text="Tempo de Expiração (dias):").pack(pady=(10, 2))
        ctk.CTkEntry(self, textvariable=self.expiration, width=500).pack()

        self.exec_button = ctk.CTkButton(self, text="Executar", command=self.executar)
        self.exec_button.pack(pady=20)

        self.progress = ctk.CTkProgressBar(self, width=500)
        self.progress.set(0)
        self.progress.pack(pady=10)

        self.log_area = ctk.CTkTextbox(self, width=700, height=250)
        self.log_area.pack(pady=10, padx=20, fill="both", expand=True)

    def selecionar_arquivo(self, var_path):
        caminho = fd.askopenfilename(filetypes=[("Planilhas Excel", "*.xlsx")])
        if caminho:
            var_path.set(caminho)

    def log(self, msg):
        self.log_area.insert("end", msg + "\n")
        self.log_area.see("end")
        # Log em arquivo
        os.makedirs("logs", exist_ok=True)
        with open(f"logs/log_{datetime.now().strftime('%Y-%m-%d')}.txt", "a", encoding="utf-8") as f:
            f.write(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}\n")

    def executar(self):
        if not all([self.emails_enviados_path.get(), self.pesquisa_path.get(), self.token.get(), self.expiration.get()]):
            mb.showerror("Erro", "Todos os campos são obrigatórios!")
            return

        self.logs_excel.clear()
        self.progress.set(0)
        self.exec_button.configure(state="disabled", text="Processando...")
        threading.Thread(target=self.processar_planilha, daemon=True).start()

    def processar_planilha(self):
        try:
            # Validações iniciais
            try:
                expiration_int = int(self.expiration.get())
                if expiration_int <= 0:
                    raise ValueError()
            except ValueError:
                self.log(f"❌ Erro: O tempo de expiração deve ser um número inteiro maior que zero.")
                return

            # Carrega e-mails já enviados
            try:
                wb_enviados = load_workbook(self.emails_enviados_path.get())
                sheet_enviados = wb_enviados.active
                emails_pesquisados = {str(row[0]).strip().lower() for row in sheet_enviados.iter_rows(min_row=2, values_only=True) if row and row[0]}
                self.log(f"🔎 {len(emails_pesquisados)} e-mails encontrados na base de controle.")
            except Exception as e:
                self.log(f"❌ Erro ao ler a planilha de e-mails enviados: {e}")
                return

            # Carrega dados da pesquisa
            try:
                wb_pesquisa = load_workbook(self.pesquisa_path.get())
                sheet_pesquisa = wb_pesquisa.active
                total = sheet_pesquisa.max_row - 1
                if total <= 0:
                    self.log("⚠️ A planilha de pesquisa está vazia ou contém apenas o cabeçalho.")
                    return
            except Exception as e:
                self.log(f"❌ Erro ao ler a planilha da pesquisa: {e}")
                return

            headers = {"Authorization": f"Bearer {self.token.get()}", "Content-Type": "application/json"}
            URL_API = "https://api.amplifique.me/partners/cf"

            for i, row in enumerate(sheet_pesquisa.iter_rows(min_row=2, values_only=True), start=1):
                self.progress.set(i / total)

                if not (row and len(row) >= 10):
                    msg = f"⚠️ Linha {i+1} ignorada por ter menos de 10 colunas."
                    self.log(msg)
                    self.logs_excel.append(["", "Ignorado", "", msg])
                    continue

                name, email, company, customer_id, transaction_id, unidade_negocio, empresa, filial, celula_de_atendimento, vip = row[:10]
                campos = [name, email, company, customer_id, transaction_id, unidade_negocio, empresa, filial, celula_de_atendimento, vip]

                if not campo_valido(email) or not email_valido(email):
                    msg = f"⚠️ E-mail inválido ou vazio na linha {i+1}: '{email}'"
                    self.log(msg)
                    self.logs_excel.append([name, "Erro", "", msg])
                    continue
                
                email_norm = str(email).strip().lower()

                if not all(map(campo_valido, campos)):
                    msg = f"⚠️ Dados incompletos na linha {i+1} para o e-mail {email}."
                    self.log(msg)
                    self.logs_excel.append([name, "Erro", "", msg])
                    continue

                if email_norm in emails_pesquisados:
                    msg = f"⏭️ E-mail já pesquisado anteriormente, ignorando: {email}"
                    self.log(msg)
                    self.logs_excel.append([name, "Já pesquisado", "", msg])
                    continue

                if "@avipam.com.br" in email_norm:
                    msg = f"⛔ E-mail interno ignorado: {email}"
                    self.log(msg)
                    self.logs_excel.append([name, "Ignorado (interno)", "", msg])
                    continue

                payload = {
                    "name": name, "email": email, "company": company, "channel": "email",
                    "customerId": customer_id, "transactionId": transaction_id,
                    "custom_fields": {
                        "unidade de negócio": unidade_negocio, "empresa": empresa, "filial": filial,
                        "celula_de_atendimento": celula_de_atendimento, "vip": vip
                    },
                    "expiration": expiration_int
                }

                try:
                    response = requests.post(URL_API, json=payload, headers=headers, verify=False, timeout=15)
                    if response.status_code == 200:
                        retorno_id = response.json().get("_id", "N/A")
                        msg = f"✅ Enviado para {name} ({email}) - ID: {retorno_id}"
                        self.logs_excel.append([name, "Sucesso", retorno_id, msg])
                    else:
                        msg = f"❌ Erro ao enviar para {name} ({email}) - Status: {response.status_code} - {response.text}"
                        self.logs_excel.append([name, "Erro API", "", msg])
                    self.log(msg)
                except requests.exceptions.RequestException as e:
                    msg = f"❌ Falha de conexão ao enviar para {name}: {e}"
                    self.log(msg)
                    self.logs_excel.append([name, "Erro Conexão", "", msg])

            self.exportar_log_excel()

        except Exception as e:
            self.log(f"💥 Ocorreu um erro fatal no processamento: {e}")
        finally:
            self.exec_button.configure(state="normal", text="Executar")
            mb.showinfo("Concluído", "Processamento finalizado! Verifique os logs para mais detalhes.")

    def exportar_log_excel(self):
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Log de Envio"
            ws.append(["Nome", "Status", "ID Retorno", "Mensagem"])
            for log_entry in self.logs_excel:
                ws.append(log_entry)
            
            timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            log_filename = f"log_de_envio_{timestamp}.xlsx"
            wb.save(log_filename)
            self.log(f"📁 Log detalhado exportado para: {log_filename}")
        except Exception as e:
            self.log(f"❌ Erro ao exportar o log para Excel: {e}")

if __name__ == "__main__":
    splash = SplashScreen()
    splash.mainloop()
    
